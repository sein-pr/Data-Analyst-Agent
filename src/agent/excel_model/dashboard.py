from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..quickchart_client import QuickChartClient
from ..logger import get_logger

logger = get_logger(__name__)


def build_dashboard_sheet(
    workbook,
    sheet_name: str,
    department: str,
    kpi_items: List[tuple[str, str]],
    kpi_descriptions: Dict[str, str] | None,
    visuals: List[Dict[str, object]],
    nav_items: List[str],
    tables: List[Dict[str, object]] | None = None,
) -> None:
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        workbook.remove(ws)
    ws = workbook.create_sheet(sheet_name, 0)

    # Layout sizing
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 3
    for col in range(3, 13):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    for r in range(3, 8):
        ws.row_dimensions[r].height = 20
    for r in range(8, 12):
        ws.row_dimensions[r].height = 18
    for r in range(13, 25):
        ws.row_dimensions[r].height = 18
    for r in range(25, 34):
        ws.row_dimensions[r].height = 18

    # Colors
    header_fill = PatternFill("solid", fgColor="0B3B4C")
    nav_fill = PatternFill("solid", fgColor="0B1F2A")
    nav_active_fill = PatternFill("solid", fgColor="135A63")
    card_fill = PatternFill("solid", fgColor="111827")
    card_accent_fill = PatternFill("solid", fgColor="1F2937")
    panel_fill = PatternFill("solid", fgColor="E2E8F0")
    footer_fill = PatternFill("solid", fgColor="0B3B4C")
    white_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="0F172A")

    # Header (avoid writing to non-anchor merged cells)
    ws.merge_cells("C1:J2")
    ws.merge_cells("K1:L2")
    ws["C1"] = f"{department.title()} Dashboard"
    ws["C1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["C1"].alignment = Alignment(vertical="center", horizontal="left")
    ws["K1"] = datetime.utcnow().strftime("%Y-%m-%d")
    ws["K1"].font = Font(color="FFFFFF", bold=True)
    ws["K1"].alignment = Alignment(vertical="center", horizontal="right")
    for cell in ws["C1:J2"][0]:
        cell.fill = header_fill
    for cell in ws["C1:J2"][1]:
        cell.fill = header_fill
    for cell in ws["K1:L2"][0]:
        cell.fill = header_fill
    for cell in ws["K1:L2"][1]:
        cell.fill = header_fill

    # Remove gridlines for a cleaner dashboard look
    ws.sheet_view.showGridLines = False

    # Left nav
    ws.merge_cells("A1:B2")
    ws["A1"] = "Navigation"
    ws["A1"].font = white_font
    ws["A1"].fill = nav_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    row = 3
    for item in nav_items:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = item
        ws[f"A{row}"].font = Font(color="FFFFFF")
        ws[f"A{row}"].fill = nav_active_fill if item.endswith(department.title()) else nav_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        if item in workbook.sheetnames:
            ws[f"A{row}"].hyperlink = f"#{item}!A1"
        row += 1

    # KPI cards (2 rows x 3 cols)
    start_row = 3
    start_col = 3
    card_width = 3
    card_height = 2
    max_cards = 6
    border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )

    for idx, (label, value) in enumerate(kpi_items[:max_cards]):
        r = start_row + (idx // 3) * (card_height + 1)
        c = start_col + (idx % 3) * (card_width + 1)
        cell_range = f"{get_column_letter(c)}{r}:{get_column_letter(c+card_width-1)}{r+card_height-1}"
        ws.merge_cells(cell_range)
        cell = ws[f"{get_column_letter(c)}{r}"]
        cell.value = f"{label}\n{value}"
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = Font(size=11, bold=True, color="FFFFFF")
        for row_cells in ws[cell_range]:
            for cell_item in row_cells:
                cell_item.fill = card_fill
                cell_item.border = border
        # accent strip
        accent_cell = ws[f"{get_column_letter(c)}{r}"]
        accent_cell.fill = card_accent_fill
        # icon image
        icon_key = _kpi_icon_key(label)
        icon_path = _get_icon_path(icon_key)
        if icon_path:
            img = XLImage(str(icon_path))
            img.width = 20
            img.height = 20
            ws.add_image(img, f"{get_column_letter(c)}{r}")

    # KPI descriptions (light panel, dark text)
    if kpi_descriptions:
        desc_start = 8
        ws.merge_cells("C8:L8")
        ws["C8"] = "KPI Highlights"
        ws["C8"].font = Font(bold=True, size=12, color="0F172A")
        ws["C8"].alignment = Alignment(vertical="center", horizontal="left")
        for row_cells in ws["C8:L11"]:
            for cell_item in row_cells:
                cell_item.fill = panel_fill
        for idx, (label, _) in enumerate(kpi_items[:3], start=1):
            row_idx = desc_start + idx
            ws.merge_cells(f"C{row_idx}:L{row_idx}")
            ws[f"C{row_idx}"] = f"{label}: {kpi_descriptions.get(label, '')}"
            ws[f"C{row_idx}"].alignment = Alignment(wrap_text=True)
            ws[f"C{row_idx}"].font = Font(color="0F172A", size=10)

    # Charts area
    chart_row = 13
    chart_col_left = 3
    chart_col_right = 9

    qc = QuickChartClient()
    for idx, visual in enumerate(visuals[:2]):
        image_bytes = _render_quickchart(visual, qc)
        if not image_bytes:
            continue
        chart_col = chart_col_left if idx == 0 else chart_col_right
        anchor = f"{get_column_letter(chart_col)}{chart_row}"
        image_path = Path(f"state/tmp_chart_{sheet_name}_{idx}.png")
        image_path.parent.mkdir(parents=True, exist_ok=True)
        image_path.write_bytes(image_bytes)
        img = XLImage(str(image_path))
        img.width = 360
        img.height = 220
        ws.add_image(img, anchor)

    # Tables area
    if tables:
        table_row = 25
        table_col_left = 3
        table_col_right = 9
        for idx, table in enumerate(tables[:2]):
            col = table_col_left if idx == 0 else table_col_right
            _render_table(
                ws,
                start_row=table_row,
                start_col=col,
                title=table.get("title", "Table"),
                headers=table.get("headers", []),
                rows=table.get("rows", []),
            )

    # Footer
    ws.merge_cells("C34:L35")
    ws["C34"] = "Generated by Autonomous Data Analyst Agent"
    ws["C34"].font = white_font
    ws["C34"].alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws["C34:L35"][0]:
        cell.fill = footer_fill


def _render_table(ws, start_row: int, start_col: int, title: str, headers, rows) -> None:
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3B4C")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )
    header_row = start_row + 1
    for idx, header in enumerate(headers, start=0):
        cell = ws.cell(row=header_row, column=start_col + idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=0):
            cell = ws.cell(row=header_row + r_idx, column=start_col + c_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            cell.font = Font(color="E2E8F0")


def _kpi_icon_key(label: str) -> str:
    text = label.lower()
    if "revenue" in text:
        return "revenue"
    if "margin" in text:
        return "margin"
    if "cost" in text:
        return "cost"
    if "discount" in text:
        return "discount"
    if "units" in text:
        return "units"
    if "growth" in text:
        return "growth"
    return "default"


def _get_icon_path(key: str):
    try:
        from PIL import Image, ImageDraw
    except Exception:
        return None

    icon_dir = Path("state/excel_icons")
    icon_dir.mkdir(parents=True, exist_ok=True)
    icon_path = icon_dir / f"{key}.png"
    if icon_path.exists():
        return icon_path

    color_map = {
        "revenue": "#22C55E",
        "margin": "#38BDF8",
        "cost": "#F97316",
        "discount": "#F59E0B",
        "units": "#A78BFA",
        "growth": "#10B981",
        "default": "#64748B",
    }
    img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    fill = color_map.get(key, "#64748B")
    draw.ellipse((4, 4, 60, 60), fill=fill)
    img.save(icon_path)
    return icon_path


def _render_quickchart(visual: Dict[str, object], qc: QuickChartClient) -> bytes | None:
    labels = []
    values = []
    for item in visual.get("data", []):
        if not isinstance(item, dict):
            continue
        labels.append(str(item.get(visual.get("x", ""), "")))
        raw = str(item.get(visual.get("y", ""), "0")).replace(",", "")
        try:
            values.append(float(raw))
        except ValueError:
            values.append(0.0)
    if not labels:
        return None
    chart_type = visual.get("type", "bar")
    config = {
        "type": "bar" if chart_type == "donut" else chart_type,
        "data": {"labels": labels, "datasets": [{"data": values, "backgroundColor": "#38BDF8"}]},
        "options": {
            "plugins": {"legend": {"display": False}},
            "scales": {
                "x": {"ticks": {"color": "#E2E8F0"}, "grid": {"color": "#1F2937"}},
                "y": {"ticks": {"color": "#E2E8F0"}, "grid": {"color": "#1F2937"}},
            },
        },
    }
    return qc.render_chart(config, width=800, height=300, background="#0B1F2A")
