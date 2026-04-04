from __future__ import annotations

from datetime import datetime

from openpyxl.styles import Alignment, Font, PatternFill


def build_dashboard_home(workbook, departments: list[str]) -> None:
    sheet_name = "Dashboard Home"
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        workbook.remove(ws)
    ws = workbook.create_sheet(sheet_name, 0)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    header_fill = PatternFill("solid", fgColor="0F766E")
    ws.merge_cells("A1:B2")
    ws["A1"] = "Department Selector"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws["A1:B2"][0]:
        cell.fill = header_fill

    ws["A4"] = "Choose a dashboard:"
    ws["A4"].font = Font(size=12, bold=True)

    row = 6
    for dept in departments:
        label = f"Dashboard - {dept.title()}"
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(size=12, bold=True, color="0F172A")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        if label in workbook.sheetnames:
            ws[f"A{row}"].hyperlink = f"#{label}!A1"
        row += 2

    ws["A12"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d')}"
    ws["A12"].font = Font(size=10, color="475569")
