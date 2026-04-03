from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass
class ExcelModel:
    workbook: object
    sheet: object
    path: Path
    name: str


def load_model(model_path: str, sheet_name: Optional[str] = None) -> ExcelModel:
    path = Path(model_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel model not found: {model_path}")
    wb = load_workbook(path, data_only=False)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}")
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    return ExcelModel(workbook=wb, sheet=sheet, path=path, name=path.name)
